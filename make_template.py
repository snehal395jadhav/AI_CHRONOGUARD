"""
Run once:  python make_template.py
Produces:  Calendar_QC_Report_Template.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from openpyxl.drawing.image import Image as XLImage
from datetime import date
import os

OUT = os.path.join(os.path.dirname(__file__), "Calendar_QC_Report_Template.xlsx")

wb = Workbook()

# ── palette ──────────────────────────────────────────────────────────────
NAVY     = "0B2258"
BLUE     = "1565C0"
LBLUE    = "4DA3FF"
WHITE    = "FFFFFF"
LGRAY    = "F4F7FB"
MGRAY    = "D0D8E8"
DGRAY    = "6C757D"

C_PASS   = "C6EFCE"; FC_PASS  = "276221"
C_FAIL   = "FFC7CE"; FC_FAIL  = "9C0006"
C_WARN   = "FFEB9C"; FC_WARN  = "9C6500"
C_NA     = "E9ECEF"; FC_NA    = "6C757D"
C_HDR    = NAVY;     FC_HDR   = WHITE

def fill(hex6):       return PatternFill("solid", fgColor=hex6)
def font(hex6, sz=10, bold=False, italic=False):
    return Font(color=hex6, size=sz, bold=bold, italic=italic, name="Segoe UI")
def ctr():            return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(wrap=False): return Alignment(horizontal="left",   vertical="center", wrap_text=wrap)
def thin_border(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def thick_border(color=NAVY):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

BDR   = thin_border()
BDR_T = thick_border()

def hdr_cell(ws, row, col, text, bg=C_HDR, fg=FC_HDR, sz=10, colspan=1):
    cell = ws.cell(row, col, text)
    cell.fill      = fill(bg)
    cell.font      = font(fg, sz, bold=True)
    cell.alignment = ctr()
    cell.border    = BDR_T
    if colspan > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col+colspan-1)
    return cell

def auto_width(ws, overrides=None):
    overrides = overrides or {}
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        if cl in overrides:
            ws.column_dimensions[cl].width = overrides[cl]
            continue
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[cl].width = min(max(w + 2, 10), 40)

def dv_list(ws, formula, sqref):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                        showDropDown=False, showErrorMessage=True,
                        errorTitle="Invalid", error="Choose from the list.")
    ws.add_data_validation(dv)
    dv.sqref = sqref
    return dv

STATUS_LIST  = '"PASS,FAIL,WARN,N/A"'
DESIGN_LIST  = '"Year-View 36x24,Portrait 22x29,Portrait 12x17,Landscape 11x8,Three-Month 12x27,Deskpad 18x11,Deskpad 22x17,Other"'
SEV_LIST     = '"CRITICAL,HIGH,MEDIUM,LOW"'
RESOLVED_LIST= '"Open,In Progress,Fixed,Verified"'

# ════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER  (company info + summary)
# ════════════════════════════════════════════════════════════════════════
ws_cov = wb.active
ws_cov.title = "Cover"
ws_cov.sheet_view.showGridLines = False

# banner
for row in range(1, 6):
    for col in range(1, 14):
        c = ws_cov.cell(row, col, "")
        c.fill = fill(NAVY)
ws_cov.row_dimensions[1].height = 14
ws_cov.row_dimensions[2].height = 14
ws_cov.row_dimensions[3].height = 22
ws_cov.row_dimensions[4].height = 14
ws_cov.row_dimensions[5].height = 14

title_cell = ws_cov.cell(3, 1,
    "  NAVNEET AI ChronoGuard  -  Calendar QC Audit Report")
title_cell.font      = font(WHITE, 18, bold=True)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
title_cell.fill      = fill(NAVY)
ws_cov.merge_cells("A1:M5")

# sub-banner
ws_cov.merge_cells("A6:M6")
sub = ws_cov.cell(6, 1, "  Pre-Press Quality Control — Zero Defect Standard")
sub.fill      = fill(BLUE)
sub.font      = font(WHITE, 11, italic=True)
sub.alignment = Alignment(horizontal="left", vertical="center")
ws_cov.row_dimensions[6].height = 20

# ── Info block ──────────────────────────────────────────────────────────
LABELS = [
    (8,  "Product / SKU Code",   "B8",  "F8"),
    (9,  "Calendar Name / Title","B9",  "F9"),
    (10, "Calendar Year",        "B10", "D10"),
    (11, "Design Type",          "B11", "D11"),
    (12, "No. of Pages (PDF)",   "B12", "D12"),
    (14, "Auditor Name",         "B14", "F14"),
    (15, "Audit Date",           "B15", "D15"),
    (16, "Company / Client",     "B16", "F16"),
    (17, "Project Reference",    "B17", "F17"),
]
for row, lbl, val_s, val_e in LABELS:
    ws_cov.row_dimensions[row].height = 20
    lc = ws_cov.cell(row, 2, lbl)
    lc.font      = font(NAVY, 10, bold=True)
    lc.alignment = left()
    # value cell (merged)
    sc = val_s.replace(lbl[0],"")  # just get col letters
    sr, sc2 = int(val_s[1:]), val_s[0]
    er, ec2 = int(val_e[1:]), val_e[0]
    ws_cov.merge_cells(f"{val_s}:{val_e}")
    vc = ws_cov[val_s]
    vc.fill      = fill(LGRAY)
    vc.border    = BDR
    vc.alignment = left()
    vc.font      = font(NAVY, 10)

# Design dropdown on cover
dv_list(ws_cov, DESIGN_LIST, "B11:D11")

# ── Overall status box ───────────────────────────────────────────────────
ws_cov.row_dimensions[19].height = 16
ws_cov.row_dimensions[20].height = 32
ws_cov.row_dimensions[21].height = 16

ws_cov.merge_cells("B19:H19")
ot = ws_cov.cell(19, 2, "OVERALL AUDIT STATUS")
ot.font = font(NAVY, 11, bold=True); ot.alignment = ctr()

ws_cov.merge_cells("B20:H20")
os_ = ws_cov.cell(20, 2, "")          # auditor fills this in
os_.fill      = fill(C_NA)
os_.border    = BDR_T
os_.alignment = ctr()
os_.font      = font(DGRAY, 14, bold=True)
dv_list(ws_cov, STATUS_LIST, "B20:H20")

ws_cov.merge_cells("I20:M20")
hint = ws_cov.cell(20, 9, "← Select: PASS / FAIL / WARN / N/A")
hint.font = font(DGRAY, 9, italic=True); hint.alignment = left()

# ── Legend ───────────────────────────────────────────────────────────────
ws_cov.row_dimensions[23].height = 16
ws_cov.merge_cells("B23:M23")
lh = ws_cov.cell(23, 2, "STATUS COLOR LEGEND")
lh.font = font(NAVY, 10, bold=True)

legend = [("PASS","All checks passed — ready to print.", C_PASS, FC_PASS),
          ("FAIL","One or more critical errors found.",    C_FAIL, FC_FAIL),
          ("WARN","Minor issue — review recommended.",     C_WARN, FC_WARN),
          ("N/A", "Not applicable / page skipped.",        C_NA,   FC_NA)]
for i, (s, desc, bg, fg) in enumerate(legend):
    row = 24 + i
    ws_cov.row_dimensions[row].height = 18
    sc = ws_cov.cell(row, 2, s)
    sc.fill = fill(bg); sc.font = font(fg, 10, bold=True); sc.border = BDR; sc.alignment = ctr()
    dc = ws_cov.cell(row, 3, desc)
    dc.fill = fill(LGRAY); dc.font = font(NAVY, 10); dc.alignment = left(); dc.border = BDR
    ws_cov.merge_cells(f"C{row}:M{row}")

# col widths for cover
for c in range(1,14):
    ws_cov.column_dimensions[get_column_letter(c)].width = (3 if c==1 else 16)

ws_cov.column_dimensions["A"].width = 2
ws_cov.column_dimensions["B"].width = 22

# ════════════════════════════════════════════════════════════════════════
# SHEET 2 — QC CHECKLIST  (one row per calendar month/page)
# ════════════════════════════════════════════════════════════════════════
ws_qc = wb.create_sheet("QC Checklist")
ws_qc.sheet_view.showGridLines = False
ws_qc.freeze_panes = "D4"

# ── Section banner ───────────────────────────────────────────────────────
ws_qc.merge_cells("A1:V1")
b1 = ws_qc.cell(1, 1, "  NAVNEET AI ChronoGuard  ·  QC Checklist — Per Month / Per Page")
b1.fill = fill(NAVY); b1.font = font(WHITE, 12, bold=True)
b1.alignment = Alignment(horizontal="left", vertical="center")
ws_qc.row_dimensions[1].height = 24

ws_qc.merge_cells("A2:V2")
b2 = ws_qc.cell(2, 1, "  Fill one row per calendar page. Use dropdown selectors for all Status columns.")
b2.fill = fill(BLUE); b2.font = font(WHITE, 9, italic=True)
b2.alignment = Alignment(horizontal="left", vertical="center")
ws_qc.row_dimensions[2].height = 16

# ── Column definitions ───────────────────────────────────────────────────
# Group A: Page Info (cols 1-6)
# Group B: Main Checks (cols 7-13)
# Group C: Mini Grid Checks (cols 14-21)
# Group D: Result + Notes (cols 22-24)

GROUP_HDRS = [
    (1,  6,  "PAGE INFORMATION",     "1565C0"),
    (7,  13, "MAIN CALENDAR CHECKS (7)",  NAVY),
    (14, 21, "MINI REFERENCE STRIP CHECKS (8)",  "0A5C36"),
    (22, 24, "RESULT",               "6F42C1"),
]
ws_qc.row_dimensions[3].height = 22
for sc, ec, lbl, bg in GROUP_HDRS:
    if sc == ec:
        hdr_cell(ws_qc, 3, sc, lbl, bg)
    else:
        hdr_cell(ws_qc, 3, sc, lbl, bg, colspan=ec-sc+1)

COL_HDRS = [
    # Page Info
    "Page #", "Month", "Year", "Design Type", "Product Code", "Calendar Title",
    # Main checks
    "1. Leap Year", "2. Sequential", "3. Misplacement",
    "4. Alignment", "5. Slash / Stacked", "6. Non-Bold Ovf", "7. Spelling",
    # Mini Grid
    "MG-1 Detection", "MG-2 Completeness", "MG-3 Column",
    "MG-4 First Day", "MG-5 Last Day", "MG-6 Leap Year",
    "MG-7 Slash", "MG-8 Continuity",
    # Result
    "OVERALL STATUS", "Holiday Check", "Comments / Issues Found",
]
ws_qc.row_dimensions[4].height = 42
for col, h in enumerate(COL_HDRS, 1):
    bg = (C_HDR if col <= 6 else
          "1A2A6C" if col <= 13 else
          "0A5C36" if col <= 21 else
          "4A148C")
    c = ws_qc.cell(4, col, h)
    c.fill      = fill(bg)
    c.font      = font(WHITE, 9, bold=True)
    c.alignment = ctr()
    c.border    = BDR

# ── Data rows (30 rows pre-formatted) ────────────────────────────────────
STATUS_COLS_MAIN = list(range(7, 14))    # cols 7-13
STATUS_COLS_MINI = list(range(14, 22))   # cols 14-21
STATUS_COLS_OV   = [22, 23]             # overall + holiday

for row in range(5, 36):
    h = 20
    ws_qc.row_dimensions[row].height = h
    for col in range(1, 25):
        c = ws_qc.cell(row, col, "")
        c.border    = BDR
        c.alignment = ctr() if col != 24 else left(wrap=True)
        # Alternate row fill for readability
        if row % 2 == 0:
            c.fill = fill(LGRAY)
        else:
            c.fill = fill(WHITE)
        c.font = font(NAVY, 10)
    # Bigger comments column
    ws_qc.row_dimensions[row].height = 22

# ── Data Validations ─────────────────────────────────────────────────────
DATA_ROWS = "5:35"
dv_list(ws_qc, DESIGN_LIST,  f"D{DATA_ROWS[0]}:D35")

for col in STATUS_COLS_MAIN + STATUS_COLS_MINI + STATUS_COLS_OV:
    cl = get_column_letter(col)
    dv_list(ws_qc, STATUS_LIST, f"{cl}5:{cl}35")

# ── Column widths ─────────────────────────────────────────────────────────
CW = {
    "A":7, "B":13, "C":7, "D":20, "E":16, "F":20,          # page info
    "G":11,"H":11, "I":12,"J":11, "K":13,  "L":12,"M":11,  # main checks
    "N":12,"O":13, "P":11,"Q":12, "R":11,  "S":11,"T":10,"U":12, # mini
    "V":14,"W":13, "X":34,                                   # result
}
for cl, w in CW.items():
    ws_qc.column_dimensions[cl].width = w

# ── Conditional fill helper comment ───────────────────────────────────────
note_cell = ws_qc.cell(37, 1,
    "TIP: After filling status values, use Ctrl+A → Format Cells → "
    "or apply conditional formatting rules for automatic color coding.")
note_cell.font      = font(DGRAY, 8, italic=True)
note_cell.alignment = left()
ws_qc.merge_cells("A37:X37")

# ════════════════════════════════════════════════════════════════════════
# SHEET 3 — ISSUES LOG  (detailed per-issue tracking)
# ════════════════════════════════════════════════════════════════════════
ws_log = wb.create_sheet("Issues Log")
ws_log.sheet_view.showGridLines = False
ws_log.freeze_panes = "A5"

ws_log.merge_cells("A1:L1")
bl = ws_log.cell(1, 1, "  NAVNEET AI ChronoGuard  ·  Issues Log — Detailed Error Tracking")
bl.fill = fill(NAVY); bl.font = font(WHITE, 12, bold=True)
bl.alignment = Alignment(horizontal="left", vertical="center")
ws_log.row_dimensions[1].height = 24

ws_log.merge_cells("A2:L2")
bl2 = ws_log.cell(2, 1, "  Document every FAIL / WARN here. One row per issue.")
bl2.fill = fill("9C0006"); bl2.font = font(WHITE, 9, italic=True)
bl2.alignment = Alignment(horizontal="left", vertical="center")
ws_log.row_dimensions[2].height = 16

LOG_HDRS = [
    "Issue #", "Page #", "Month / Year", "Design Type",
    "Checker", "Check Name", "Mini Month\n(if Mini Grid)",
    "Description of Error / Issue Found",
    "Severity", "Action Required", "Responsible", "Status"
]
CHECKER_LIST  = '"Date Checker,Mini Grid Checker,Holiday Checker,Visual Review,Other"'
CHECK_LIST    = ('"1. Leap Year,2. Sequential,3. Misplacement,4. Alignment,'
                 '5. Slash/Stacked,6. Non-Bold Ovf,7. Spelling,'
                 'MG-1 Detection,MG-2 Completeness,MG-3 Column Placement,'
                 'MG-4 First Day,MG-5 Last Day,MG-6 Leap Year,'
                 'MG-7 Slash,MG-8 Continuity,Holiday Missing,'
                 'Holiday Wrong Date,Holiday Spelling,Other"')

ws_log.row_dimensions[4].height = 36
for col, h in enumerate(LOG_HDRS, 1):
    c = ws_log.cell(4, col, h)
    c.fill      = fill(C_HDR)
    c.font      = font(WHITE, 10, bold=True)
    c.alignment = ctr()
    c.border    = BDR

for row in range(5, 55):
    ws_log.row_dimensions[row].height = 22
    for col in range(1, 13):
        c = ws_log.cell(row, col, "")
        c.border    = BDR
        c.alignment = ctr() if col != 8 and col != 10 else left(wrap=True)
        c.fill      = fill(LGRAY if row % 2 == 0 else WHITE)
        c.font      = font(NAVY, 10)
    # Auto-number issue # column
    ws_log.cell(row, 1, row - 4)
    ws_log.cell(row, 1).font = font(DGRAY, 10)

dv_list(ws_log, DESIGN_LIST,   "D5:D54")
dv_list(ws_log, CHECKER_LIST,  "E5:E54")
dv_list(ws_log, CHECK_LIST,    "F5:F54")
dv_list(ws_log, SEV_LIST,      "I5:I54")
dv_list(ws_log, RESOLVED_LIST, "L5:L54")

MONTH_LIST = '"January,February,March,April,May,June,July,August,September,October,November,December,All Months"'
dv_list(ws_log, MONTH_LIST, "G5:G54")

LOG_CW = {"A":9,"B":8,"C":18,"D":20,"E":16,"F":22,"G":14,
          "H":42,"I":12,"J":30,"K":18,"L":15}
for cl, w in LOG_CW.items():
    ws_log.column_dimensions[cl].width = w

# ════════════════════════════════════════════════════════════════════════
# SHEET 4 — HOLIDAY CHECK  (holiday-specific validation)
# ════════════════════════════════════════════════════════════════════════
ws_hol = wb.create_sheet("Holiday Check")
ws_hol.sheet_view.showGridLines = False
ws_hol.freeze_panes = "A5"

ws_hol.merge_cells("A1:J1")
bh = ws_hol.cell(1, 1, "  NAVNEET AI ChronoGuard  ·  Holiday Check Report")
bh.fill = fill("0A5C36"); bh.font = font(WHITE, 12, bold=True)
bh.alignment = Alignment(horizontal="left", vertical="center")
ws_hol.row_dimensions[1].height = 24

ws_hol.merge_cells("A2:J2")
bh2 = ws_hol.cell(2, 1, "  One row per expected holiday. Mark Status after comparing PDF vs. reference list.")
bh2.fill = fill("1A7C4A"); bh2.font = font(WHITE, 9, italic=True)
bh2.alignment = Alignment(horizontal="left", vertical="center")
ws_hol.row_dimensions[2].height = 16

HOL_HDRS = [
    "#", "Holiday Name", "Expected Date\n(dd/mm/yyyy)",
    "Expected Day\n(Mon, Tue …)", "Month", "Page #",
    "Text Found in PDF", "Status", "Notes / Action", "Verified By"
]
HOL_STATUS = ('"✅ Exact Match,⚠️ Wrong Date,⚠️ Spelling Error,'
              '❌ Not Found,🔴 Misplaced – Wrong Month,'
              '🆕 Extra – Not in Reference,N/A"')

ws_hol.row_dimensions[4].height = 36
for col, h in enumerate(HOL_HDRS, 1):
    c = ws_hol.cell(4, col, h)
    c.fill      = fill("0A5C36")
    c.font      = font(WHITE, 10, bold=True)
    c.alignment = ctr()
    c.border    = BDR

for row in range(5, 65):
    ws_hol.row_dimensions[row].height = 20
    for col in range(1, 11):
        c = ws_hol.cell(row, col, "")
        c.border    = BDR
        c.alignment = ctr() if col not in (7,9) else left(wrap=True)
        c.fill      = fill(LGRAY if row % 2 == 0 else WHITE)
        c.font      = font(NAVY, 10)
    ws_hol.cell(row, 1, row - 4)
    ws_hol.cell(row, 1).font = font(DGRAY, 10)

dv_list(ws_hol, HOL_STATUS,  "H5:H64")
dv_list(ws_hol, MONTH_LIST,  "E5:E64")

HOL_CW = {"A":6,"B":32,"C":18,"D":16,"E":13,"F":8,
          "G":30,"H":24,"I":36,"J":16}
for cl, w in HOL_CW.items():
    ws_hol.column_dimensions[cl].width = w

# ════════════════════════════════════════════════════════════════════════
# SHEET 5 — INSTRUCTIONS
# ════════════════════════════════════════════════════════════════════════
ws_ins = wb.create_sheet("Instructions")
ws_ins.sheet_view.showGridLines = False

ws_ins.merge_cells("A1:F1")
bi = ws_ins.cell(1, 1, "  NAVNEET AI ChronoGuard  ·  How to Use This Form")
bi.fill = fill(NAVY); bi.font = font(WHITE, 13, bold=True)
bi.alignment = Alignment(horizontal="left", vertical="center")
ws_ins.row_dimensions[1].height = 28

STEPS = [
    ("STEP 1", "Cover Sheet",
     "Fill in the product code, calendar name, year, design type, auditor name and date. "
     "Leave 'Overall Audit Status' until you have completed all checks."),
    ("STEP 2", "QC Checklist",
     "Add one row for EACH calendar page (month). "
     "For every check column select PASS / FAIL / WARN / N/A from the dropdown. "
     "Fill the 'Comments' column with a brief note for any FAIL or WARN."),
    ("STEP 3", "Issues Log",
     "For every FAIL or WARN on the QC Checklist, add a detailed row here. "
     "Describe what is wrong, how severe it is, and what action is needed."),
    ("STEP 4", "Holiday Check",
     "List every expected holiday. Compare against the printed PDF and mark each one. "
     "Attach the reference holiday list (xlsx) when sending to the Navneet team."),
    ("STEP 5", "Return the File",
     "Save the file as:  CalendarQC_<ProductCode>_<YYYY-MM-DD>.xlsx  "
     "and email to your Navneet project contact."),
]

CHECKS_INFO = [
    ("1. Leap Year Logic",      "February must show 28 days (normal year) or 29 days (leap year). 2028 is a leap year."),
    ("2. Sequential Continuity","Dates must run 1 → 2 → 3 … with no gaps or repeated numbers."),
    ("3. Date Misplacement",    "Each date must appear under the correct weekday column (Sun–Sat)."),
    ("4. Data Alignment",       "Date numbers must visually align under their weekday header."),
    ("5. Slash / Stacked Dates","Combined Sun/Mon cells (e.g. 29/30) are only valid when the first date is a Sunday."),
    ("6. Non-Bold Overflow",    "Overflow dates from the previous or next month must be printed in non-bold."),
    ("7. Spelling & Proofing",  "Month names and weekday headers must be spelled correctly."),
    ("MG-1 Detection",          "All 12 mini-month reference calendars are found in the mini strip."),
    ("MG-2 Completeness",       "Every date 1 … N is present in each mini-month grid."),
    ("MG-3 Column Placement",   "Each date falls in the correct Sun–Sat column in the mini grid."),
    ("MG-4 First-Day Offset",   "Day 1 of each mini-month is in the correct weekday column."),
    ("MG-5 Last-Day Integrity", "The last day of each mini-month is in the correct column."),
    ("MG-6 Leap Year (Mini)",   "Mini February shows 28 or 29 days correctly."),
    ("MG-7 Slash (Mini)",       "Slash/stacked cells in mini strip follow the Sunday rule."),
    ("MG-8 Cross-Month",        "All 12 mini-months form a consecutive January–December sequence."),
]

row = 3
for step, title, desc in STEPS:
    ws_ins.row_dimensions[row].height = 14
    ws_ins.row_dimensions[row+1].height = 20
    ws_ins.row_dimensions[row+2].height = 38
    ws_ins.row_dimensions[row+3].height = 10

    ws_ins.merge_cells(f"A{row+1}:F{row+1}")
    sh = ws_ins.cell(row+1, 1, f"  {step}  —  {title}")
    sh.fill = fill(BLUE); sh.font = font(WHITE, 11, bold=True)
    sh.alignment = Alignment(horizontal="left", vertical="center")

    ws_ins.merge_cells(f"A{row+2}:F{row+2}")
    dc = ws_ins.cell(row+2, 1, desc)
    dc.fill = fill(LGRAY); dc.font = font(NAVY, 10)
    dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row += 4

# Check definitions table
row += 1
ws_ins.merge_cells(f"A{row}:F{row}")
ch = ws_ins.cell(row, 1, "  CHECK DEFINITIONS")
ch.fill = fill(NAVY); ch.font = font(WHITE, 11, bold=True)
ch.alignment = Alignment(horizontal="left", vertical="center")
ws_ins.row_dimensions[row].height = 22
row += 1

for i, (name, desc) in enumerate(CHECKS_INFO):
    ws_ins.row_dimensions[row].height = 20
    nc = ws_ins.cell(row, 1, name)
    nc.fill = fill(LGRAY if i%2==0 else WHITE)
    nc.font = font(NAVY, 10, bold=True)
    nc.alignment = left()
    nc.border = BDR

    ws_ins.merge_cells(f"B{row}:F{row}")
    dc2 = ws_ins.cell(row, 2, desc)
    dc2.fill = fill(LGRAY if i%2==0 else WHITE)
    dc2.font = font(NAVY, 10)
    dc2.alignment = left(wrap=True)
    dc2.border = BDR
    row += 1

for c in "ABCDEF":
    ws_ins.column_dimensions[c].width = (28 if c=="A" else 16)
ws_ins.column_dimensions["B"].width = 60
ws_ins.column_dimensions["F"].width = 10

# ════════════════════════════════════════════════════════════════════════
# SHEET ORDER  +  tab colors
# ════════════════════════════════════════════════════════════════════════
tab_colors = {
    "Cover":        "0B2258",
    "QC Checklist": "1565C0",
    "Issues Log":   "9C0006",
    "Holiday Check":"0A5C36",
    "Instructions": "6F42C1",
}
for ws in wb.worksheets:
    if ws.title in tab_colors:
        ws.sheet_properties.tabColor = tab_colors[ws.title]

wb.active = wb["Cover"]

# ── save ─────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Template saved: {OUT}")
